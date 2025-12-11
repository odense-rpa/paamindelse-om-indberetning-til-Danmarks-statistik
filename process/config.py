from typing import Dict, List
from openpyxl import load_workbook

def get_excel_mapping() -> Dict[str, List[str]]:
    """Henter excel-mapping"""
    global excel_mappings
    if not excel_mappings:
        raise ValueError("excel-mapping er ikke indlæst, brug load_excel_mapping først")
    return excel_mappings


def load_excel_mapping(file_path: str):    
    global excel_mappings
    try:
        # Load workbook and get first worksheet
        workbook = load_workbook(file_path)
        worksheet = workbook.active

        if worksheet is None:
            raise ValueError("Worksheet could not be loaded")

        # Initialize mapping dictionary
        mapping: Dict[str, List[str]] = {}

        # Get header row to identify POF columns (row 1)
        header_row = worksheet[1]
        headers = []
        for cell in header_row:
            if cell.value and str(cell.value).strip():
                headers.append((str(cell.value).strip(), cell.column))

        # Process each column to build mapping
        for header, col_idx in headers:
            items = []
            # Start from row 2 (skip header), process column by column
            for row in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                cell_value = row[0].value
                if cell_value and str(cell_value).strip():  # Non-blank cells only
                    items.append(str(cell_value).strip())
            mapping[header] = items

        excel_mappings = mapping

    except Exception as e:
        raise RuntimeError(
            f"Failed to load mapping from Excel file '{file_path}': {str(e)}"
        ) from e